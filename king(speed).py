from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import soundfile
import speech_recognition as sr
import os
import random
import urllib
import pandas as pd
import numpy as np
import certifi
import math
from tkinter import *
from tkinter import messagebox
from datetime import datetime, timedelta
import ssl
import csv 
import openpyxl
from twocaptcha  import TwoCaptcha

ssl._create_default_https_context = ssl._create_unverified_context

class Window(Frame):

    def __init__(self, master=None):
        Frame.__init__(self, master)        
        self.master = master
        self.pack(fill=BOTH, expand=1)

        label_frame3 = LabelFrame(self, text='Input Search Keyword', height=110, width=200)
        label_frame3.place(x=15,y=10)
        self.name_count = Entry(self, bd=1)
        self.name_count.place(x=90,y=30,width=110)
        text6 = Label(self, text="How many?")
        text6.place(x=20,y=30)

        self.city_name = Entry(self, bd=1)
        self.city_name.place(x=90,y=60,width=110)
        text6 = Label(self, text="City name?")
        text6.place(x=20,y=60)
        
        
        # create button, link it to clickExitButton()
        start_button = Button(self, text="Start", command=self.clickStartButton)
        start_button.place(x=20, y=85, width=80)
        close_button = Button(self, text="close", command=self.clickContinueButton)
        close_button.place(x=130, y=85, width=80)
    
   
        
    def clickStartButton(self):
        
        name_count = self.name_count.get()        
        print(name_count)

        count_to_be_created = int(name_count)

        excel = openpyxl.load_workbook('surname.xlsx')
        sheet = excel.active
        excel_file = 'surname.xlsx'
        rowCount = sheet.max_row + 1
        names = []
        for row in range(1, count_to_be_created + 1):
            name = sheet.cell(row = row, column = 1).value
            names.append(name)
        print(names)

 
        # Go to site
        driver = webdriver.Chrome()
        driver.maximize_window()
        driver.get('https://www.infobel.com/en/hungary/')
        statusbar.config(text="Tool is running!")
        app.update()
        
        while True:
            try:
                c = driver.find_element("xpath", "//*[@id='sd-cmp']/div[2]/div[1]/div/div/div/div/div/div[2]/div[1]/button[3]/span")
                if c:
                    c.click()
                    break  
                time.sleep(3)
            except:
                continue
            

        # Press the company button
        companyBtn = driver.find_element("xpath", "//*[@id='search-form-header']/div[1]/span/span/span[1]")
        companyBtn.click()
        
        time.sleep(0.3)
        
        personBtn = driver.find_element("xpath", "//*[@id='search-type-select-header_listbox']/li[2]")
        personBtn.click()
        
        name_input = driver.find_element("xpath", "//*[@id='residential-search-term-input-header']")
        name_input.send_keys(names[0])
        city_input = driver.find_element("xpath", "//*[@id='search-location-input-header']")
        city_input.send_keys(self.city_name.get())
        
        
        time.sleep(0.3)

        while True:
            try:
                searchBtn = driver.find_element("xpath", "//*[@id='btn-search-header']")
                searchBtn.click()
                break       
            except:
                statusbar.config(text="Now I press searchBtn!")
                app.update() 

        while True:
            try:
                print('Solving Captcha')
                solver = TwoCaptcha('3d8e6e523cbe6f096d7acaf848672850')
                code = solver.solve_captcha(site_key='6Lf5r0IUAAAAAPUFYve7LktblGYkCuO2w3cZ4cC0', page_url='https://www.infobel.com/Landing/Abuse')
                print(f"Successfully solved the Captcha.")
                recaptcha_response_element = driver.find_element(By.ID, 'g-recaptcha-response')
                try:
                    driver.execute_script("window.___grecaptcha_cfg.clients[0].A.A.callback('{}')".format(code))
                except:
                    pass
                try:
                    driver.execute_script("window.___grecaptcha_cfg.clients[0].P.P.callback('{}')".format(code))
                except:
                    pass
                time.sleep(3)
                submit_btn = driver.find_element(By.XPATH, '//*[@id="abuse-validation-form"]/button')
                submit_btn.click()
                time.sleep(3)
                if driver.find_element("xpath", "//*[@id='residential-search-term-input-header']"):
                    break
            except:
                continue
       
        print('capture is passed, great!')
        
        for i in range(count_to_be_created + 1)[0:]:
            name_input = driver.find_element("xpath", "//*[@id='residential-search-term-input-header']")
            name_input.clear()
            try: 
                name_input.send_keys(names[i])
            except:
                break

            searchBtn = driver.find_element("xpath", "//*[@id='btn-search-header']")
            searchBtn.click()
            time.sleep(1)
            
    
            items_list = []

            # First page data collection
            try:
                numberShow = driver.find_elements(By.XPATH, "//span[text()='Display phone']")
                for span in numberShow:
                    try:
                        span.click()
                    except:
                        pass


                try:
                    items = driver.find_elements(By.XPATH, "//span[@class='detail-text' and not(text()='Website ' or text()='Display phone')]")
                    for item in items:
                        items_list.append(item.text)
                except:
                    alert = driver.switch_to.alert
                    print("Alert text: " + alert.text)
                    alert.accept()
                    print('Alert is accepted')
                    

                time.sleep(0.5)

                

                try:
                    if driver.find_element(By.XPATH, "//iframe[@title='reCAPTCHA']"):
                        while True:
                            try:
                                print('Solving Captcha')
                                solver = TwoCaptcha('db508d8f434c6565f5bf7caca459e269')
                                code = solver.solve_captcha(site_key='6Lf5r0IUAAAAAPUFYve7LktblGYkCuO2w3cZ4cC0', page_url='https://www.infobel.com/Landing/Abuse')
                                print(f"Successfully solved the Captcha.")
                                recaptcha_response_element = driver.find_element(By.ID, 'g-recaptcha-response')
                                try:
                                    driver.execute_script("window.___grecaptcha_cfg.clients[0].A.A.callback('{}')".format(code))
                                except:
                                    pass
                                try:
                                    driver.execute_script("window.___grecaptcha_cfg.clients[0].P.P.callback('{}')".format(code))
                                except:
                                    pass
                                time.sleep(3)
                                submit_btn = driver.find_element(By.XPATH, '//*[@id="abuse-validation-form"]/button')
                                submit_btn.click()
                                time.sleep(3)
                                if driver.find_element("xpath", "//*[@id='residential-search-term-input-header']"):
                                    break
                            except:
                                continue
                    
                        print('capture is passed, great!')
                except:
                    pass
            except:
                pass
            time.sleep(1)
            # main page data collection
            try:
                # ul = driver.find_element(By.XPATH, "//ul[@class='pagination']")
                # li = ul.find_elements(By.TAG_NAME, 'li')[-1]
                # a = li.find_element(By.TAG_NAME, 'a')
                while True:
                    try:
                        time.sleep(1)
                        next_page = driver.find_element(By.XPATH, "/html/body/div[3]/div[1]/div/div[4]/div[2]/div/div/ul/li[16]/a/span[1]").click()
                        # ul = driver.find_element(By.XPATH, "//ul[@class='pagination']")
                        # li = ul.find_elements(By.TAG_NAME, 'li')[-1]
                        # a = li.find_element(By.TAG_NAME, 'a')
                        # a.click()
                    except:
                        break
                    try:
                        numberShow = driver.find_elements(By.XPATH, "//span[text()='Display phone']")
                        for span in numberShow:
                            try:
                                span.click()
                            except:
                                pass

                        time.sleep(0.5)

                        try:
                            items = driver.find_elements(By.XPATH, "//span[@class='detail-text' and not(text()='Website ' or text()='Display phone')]")
                            for item in items:
                                items_list.append(item.text)
                        except:
                            alert = driver.switch_to.alert
                            print("Alert text: " + alert.text)
                            alert.accept()
                            print('Alert is accepted')
                            

                        time.sleep(0.5)

                        

                        try:
                            if driver.find_element(By.XPATH, "//iframe[@title='reCAPTCHA']"):
                                while True:
                                    try:
                                        print('Solving Captcha')
                                        solver = TwoCaptcha('db508d8f434c6565f5bf7caca459e269')
                                        code = solver.solve_captcha(site_key='6Lf5r0IUAAAAAPUFYve7LktblGYkCuO2w3cZ4cC0', page_url='https://www.infobel.com/Landing/Abuse')
                                        print(f"Successfully solved the Captcha.")
                                        recaptcha_response_element = driver.find_element(By.ID, 'g-recaptcha-response')
                                        try:
                                            driver.execute_script("window.___grecaptcha_cfg.clients[0].A.A.callback('{}')".format(code))
                                        except:
                                            pass
                                        try:
                                            driver.execute_script("window.___grecaptcha_cfg.clients[0].P.P.callback('{}')".format(code))
                                        except:
                                            pass
                                        time.sleep(3)
                                        submit_btn = driver.find_element(By.XPATH, '//*[@id="abuse-validation-form"]/button')
                                        submit_btn.click()
                                        time.sleep(3)
                                        if driver.find_element("xpath", "//*[@id='residential-search-term-input-header']"):
                                            break
                                    except:
                                        continue
                            
                                print('capture is passed, great!')
                        except:
                            pass
                    except:
                        pass
                time.sleep(1)
            except:
                pass
            statusbar.config(text="Exporting to CSV...")
            app.update()

            csv_file = str(names[i])+'.csv'        
            with open(csv_file, 'w', newline='', encoding='utf-8-sig') as file:
                writer = csv.writer(file)
                writer.writerow(['Address', 'PhoneNumber'])  # Write header row 
                for j in range(0, len(items_list), 2):
                    address = items_list[j].split()[-1]
                    phone_number = items_list[j + 1]
                    writer.writerow([address, phone_number])

            

            print(str(names[i])+'.csv: saved successfully')        

        print('All done. Ha, Ha, Wow You are genious and splendiferous!!!')
        statusbar.config(text="All done. Ha, Ha, Wow You are genious and splendiferous!!!")
        app.update()
        driver.quit()
        
    def clickContinueButton(self):
        exit()

root = Tk()
app = Window(root)
statusbar = Label(app, text="Let's go bro!", bd=1, relief=SUNKEN, anchor=E)
statusbar.pack(side=BOTTOM, fill=X)
root.wm_title("gameclub.jp")
root.geometry("230x150+1100+400")
root.resizable(False, False)
root.mainloop()